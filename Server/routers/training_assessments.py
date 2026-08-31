"""
AI-generated training assessment routes.

POST   /training/assessments/generate     — manager: generate assessment via AI
GET    /training/assessments              — manager: list assessments they created
GET    /training/assessments/mine         — new joiner: their assessments
GET    /training/assessments/{id}         — detail (with questions, no correct_answer exposed to joiner)
POST   /training/assessments/{id}/start   — new joiner starts attempt
POST   /training/assessments/{id}/submit  — new joiner submits answers → AI evaluates
GET    /training/assessments/{id}/attempts — list all attempts (manager)
GET    /training/attempts/{attempt_id}    — attempt detail with per-answer flags
"""

import io
import json
import re
from datetime import datetime, timedelta
from typing import List, Optional
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from sqlalchemy.orm import Session

from database import SessionLocal
from auth import require_role
from models import User, SmeKit, SmeKitFileV2, TrainingAssessment, TrainingQuestion, TrainingAttempt, TrainingAnswer, Notification
from schemas import (
    GenerateAssessmentRequest, TrainingAssessmentOut, TrainingQuestionOut,
    SubmitAttemptRequest, TrainingAttemptOut,
)
from config import OPENROUTER_API_KEY, MODEL_NAME

router = APIRouter(prefix="/api/training", tags=["training-assessments"])

import httpx


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ── AI helpers ───────────────────────────────────────────────────────────────

def _call_ai(prompt: str) -> str:
    headers = {
        "Authorization": f"Bearer {OPENROUTER_API_KEY}",
        "Content-Type": "application/json",
    }
    body = {
        "model": MODEL_NAME,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.3,
    }
    with httpx.Client(timeout=120) as client:
        r = client.post("https://openrouter.ai/api/v1/chat/completions", headers=headers, json=body)
        r.raise_for_status()
        return r.json()["choices"][0]["message"]["content"].strip()


def _extract_json(text: str) -> any:
    cleaned = re.sub(r"^```(?:json)?\s*", "", text.strip())
    cleaned = re.sub(r"\s*```$", "", cleaned.strip())
    return json.loads(cleaned)


def _normalize_us_quotation_marks(value):
    """Use double quotation marks for quotations without changing apostrophes.

    AI-generated editing questions sometimes wrap a quoted sentence in single
    quotation marks despite being instructed to use US English. Only paired
    quotation marks are changed; apostrophes inside words remain untouched.
    """
    if not isinstance(value, str) or "'" not in value and "‘" not in value:
        return value

    # Curly pairs are handled separately, including curly apostrophes in words.
    normalized = re.sub(
        r"‘((?:[^’\n]|(?<=\w)’(?=\w))+?)’(?=$|[\s.,!?;:)\]])",
        r'“\1”',
        value,
    )
    # Match each quoted passage separately while explicitly allowing apostrophes
    # within words (for example, "document's" and "don't").
    normalized = re.sub(
        r"(?<![\w])'((?:[^'\n]|(?<=\w)'(?=\w))+?)'(?=$|[\s.,!?;:)\]])",
        r'"\1"',
        normalized,
    )
    return normalized


def _normalize_question_quotations(question: dict) -> dict:
    """Normalize only human-readable question fields."""
    normalized = dict(question)
    normalized["question_text"] = _normalize_us_quotation_marks(
        normalized.get("question_text", "")
    )
    if isinstance(normalized.get("options"), list):
        normalized["options"] = [
            _normalize_us_quotation_marks(option)
            for option in normalized["options"]
        ]
    if isinstance(normalized.get("correct_answer"), str):
        normalized["correct_answer"] = _normalize_us_quotation_marks(
            normalized["correct_answer"]
        )
    return normalized


def _build_content_context(files: List[SmeKitFileV2]) -> str:
    parts = []
    for f in files:
        if f.transcript:
            parts.append(f"[{f.name}]\n{f.transcript[:6000]}")
        elif f.youtube_url:
            parts.append(f"[YouTube: {f.name}] URL: {f.youtube_url} (no transcript provided)")
        else:
            parts.append(f"[Document: {f.name}] (file uploaded, no text extracted)")
    return "\n\n---\n\n".join(parts) if parts else "No content available."


_EDITING_KIT_NAME_SIGNALS = (
    "eeoc", "content valid", "editing check", "editing guide",
    "style guide", "qc checklist", "content editing", "proofreading",
    "content qc", "editing qc",
)


def _is_editing_kit(kit_name: str, content: str) -> bool:
    if any(signal in (kit_name or "").lower() for signal in _EDITING_KIT_NAME_SIGNALS):
        return True
    excerpt = content[:2000].lower()
    signals = (
        "eeoc", "content validation checklist", "editing checklist",
        "sensitive keyword", "filler word", "uk english", "us english style",
    )
    return sum(1 for signal in signals if signal in excerpt) >= 2


def _generate_questions(
    content: str,
    easy_count: int, easy_type: str,
    medium_count: int, medium_type: str,
    hard_count: int, hard_type: str,
    kit_name: str = "",
    additional_instructions: str = "",
) -> List[dict]:
    total = easy_count + medium_count + hard_count
    manager_guidance = (additional_instructions or "").strip()[:2000]
    guidance_block = (
        "\nMANAGER'S ADDITIONAL INSTRUCTIONS:\n"
        f"{manager_guidance}\n"
        "Apply these instructions when choosing topics, scenarios, and wording. They must not override the required question counts, selected question types, JSON format, US English rules, or the requirement to stay grounded in the source content.\n"
        if manager_guidance else ""
    )

    def _type_label(t: str) -> str:
        return "MCQ" if t == "mcq" else "Descriptive (open-ended, no options)"

    def _type_rule(difficulty: str, t: str) -> str:
        if difficulty == "Easy":
            lang_note = " Write in plain, simple language (CEFR A2-B1 level): short sentences, common vocabulary, no jargon. Uniformly simple — context, logic, AND wording must all be easy."
        elif difficulty == "Medium":
            lang_note = " Use clear professional language (CEFR B1-B2 level): straightforward sentences, standard industry terms where needed."
        else:
            lang_note = " Language may be more complex (CEFR B2-C1 level) to match the depth of the concept being tested."

        if t == "mcq":
            return f"{'Easy' if difficulty == 'Easy' else difficulty} questions test {'direct recall of specific facts' if difficulty == 'Easy' else ('understanding of concepts' if difficulty == 'Medium' else 'analysis and application of concepts')} from the content. Use MCQ format (4 options A/B/C/D).{lang_note}"
        else:
            return f"{'Easy' if difficulty == 'Easy' else difficulty} questions test {'direct recall of specific facts' if difficulty == 'Easy' else ('application of concepts' if difficulty == 'Medium' else 'analysis and critical evaluation of concepts')} from the content. Use descriptive/open-ended format (no options).{lang_note}"

    # Build a per-difficulty type mandate that spells out EXACTLY which format each band must use.
    def _mandate_line(label: str, count: int, t: str) -> str:
        if count <= 0:
            return f"- {label}: none requested."
        fmt = "MCQ (exactly 4 options A/B/C/D, single-letter correct_answer)" if t == "mcq" else "Descriptive (options MUST be null, correct_answer starts with \"Model answer:\")"
        return f"- All {count} {label} question(s) MUST be {fmt}. No exceptions."

    type_mandate = "\n".join([
        _mandate_line("Easy", easy_count, easy_type),
        _mandate_line("Medium", medium_count, medium_type),
        _mandate_line("Hard", hard_count, hard_type),
    ])

    # Build example question objects that match the ACTUAL selected type for each band,
    # so the model mimics the correct format (it copies the example over abstract rules).
    def _example_obj(order: int, difficulty: str, t: str) -> str:
        if t == "mcq":
            return f'''    {{
      "order_index": {order},
      "difficulty": "{difficulty}",
      "question_type": "mcq",
      "question_text": "Question text here.",
      "options": ["A. First option here", "B. Second option here", "C. Third option here", "D. Fourth option here"],
      "correct_answer": "A"
    }}'''
        return f'''    {{
      "order_index": {order},
      "difficulty": "{difficulty}",
      "question_type": "descriptive",
      "question_text": "Question text here.",
      "options": null,
      "correct_answer": "Model answer: ..."
    }}'''

    example_objs = []
    ex_order = 1
    if easy_count > 0:
        example_objs.append(_example_obj(ex_order, "easy", easy_type)); ex_order += 1
    if medium_count > 0:
        example_objs.append(_example_obj(ex_order, "medium", medium_type)); ex_order += 1
    if hard_count > 0:
        example_objs.append(_example_obj(ex_order, "hard", hard_type)); ex_order += 1
    example_block = ",\n".join(example_objs)

    if _is_editing_kit(kit_name, content):
        prompt = f"""You are designing a content-validation assessment from the editing or EEOC guideline below.

[START OF CONTENT]
{content}
[END OF CONTENT]

Generate exactly {total} realistic scenario-based questions. Test whether a candidate can decide if an invented passage should be flagged under the guideline. Do not ask recall questions about what the document says or what the document is.

For every question:
1. Extract and apply a real rule from the guideline.
2. Invent a fresh two-to-four-sentence workplace passage; do not copy the guideline.
3. Use this exact question_text structure: "[passage]\\n---\\nQuestion: Should this text be flagged based on the content validation checklist? Why?"
4. For MCQs, provide exactly four comparable options labeled A–D. Each option starts with "Yes." or "No." and gives a complete explanation. Mix Yes and No options.
5. Store only A, B, C, or D in correct_answer.
6. Easy scenarios contain an obvious single violation. Medium scenarios require applying a specific rule. Hard scenarios require nuanced judgment and must include both passages that should be flagged and passages that should not.
7. Vary domains across questions. Use US English, double quotation marks, the Oxford comma, active voice, and neutral professional language.
8. NEVER use single quotation marks to enclose quoted text, passages, sentences, phrases, or answer choices. Use double quotation marks for every quotation. Use apostrophes only inside contractions and possessives.

QUESTION TYPE MANDATE:
{type_mandate}
{guidance_block}

Return only valid JSON in this structure:
{{
  "questions": [
{example_block}
  ]
}}

The questions array must contain exactly {total} items: {easy_count} Easy, {medium_count} Medium, and {hard_count} Hard."""
    else:
        prompt = f"""You are an expert trainer creating a training assessment based exclusively on the SME Kit content provided below.

Follow these rules strictly:
1. Every question must be directly answerable ONLY from the text between [START OF CONTENT] and [END OF CONTENT] below. Do not use outside knowledge or general facts.
2. Read the content first. Identify the document's actual subject. Generate questions ONLY about that subject — not about file formats, metadata, or unrelated industry practices.
3. Do not invent rules, scenarios, or facts. Every correct answer must be explicitly found word-for-word or by clear implication in the content below.
4. Use US English throughout: American spelling (-ize, -or, -er endings), double quotation marks, Oxford comma, active voice. NEVER use single quotation marks to enclose quoted text, passages, sentences, phrases, or answer choices. Use apostrophes only inside contractions and possessives.
5. {_type_rule('Easy', easy_type)}
6. {_type_rule('Medium', medium_type)}
7. {_type_rule('Hard', hard_type)}
8. For MCQ questions, vary the question formats: include both knowledge-recall questions AND error-identification questions (present a sentence with a mistake from the document and ask which option corrects it). Do not make all medium-difficulty questions identical in structure.
9. Do not repeat the same question stem pattern more than twice across the full set.
10. Keep question language professional and neutral — avoid "you" in question stems where possible.
11. MCQ options must use sentence case: capitalize only the first letter of each option text after the letter prefix (e.g., "A. The correct answer" not "A. The Correct Answer"), unless the option starts with a proper noun or technical term that is inherently capitalized. All four options must follow the same casing pattern.

ABSOLUTE PROHIBITION — these question types are forbidden and will invalidate the entire output:
- Do NOT generate any question about these instructions, prompt rules, or directives.
- Do NOT reference any phrase from these instructions in a question stem or option — phrases like "No content available", "SME Kit content", "the provided rules", "START OF CONTENT", "END OF CONTENT", "Content Editing Guidelines", "outside knowledge", or any other meta-language from this prompt.
- Do NOT generate questions about what the AI "should" or "should not" do.
- Do NOT generate questions about the document format, file type, or how the content was provided.
- If the content is too short or unclear to support a question, skip that concept — do not pad with meta-questions.

QUESTION TYPE MANDATE — this is non-negotiable and overrides any pattern you might infer from the examples below:
{type_mandate}
{guidance_block}

Before generating questions, identify: (a) the document's actual topic, and (b) four to six specific rules or concepts it covers. Base ALL questions exclusively on those concepts from the content text.

[START OF CONTENT]
{content}
[END OF CONTENT]

HARD STOP: You must generate EXACTLY {total} questions — {easy_count} Easy + {medium_count} Medium + {hard_count} Hard. Count your questions as you write them. Do NOT write question number {total + 1}. If you reach {total} questions, stop immediately.

Return ONLY valid JSON (no markdown, no explanation, no extra text before or after). The example below shows the EXACT format required for each difficulty band in this assessment — follow the question_type shown for each:
{{
  "questions": [
{example_block}
  ]
}}

Formatting rules:
- MCQ: exactly 4 options labeled A. B. C. D.; correct_answer is the single letter only (A, B, C, or D)
- Descriptive: options is null; correct_answer starts with "Model answer:"
- Order: Easy questions first (indices 1–{easy_count}), then Medium ({easy_count+1}–{easy_count+medium_count}), then Hard ({easy_count+medium_count+1}–{total})
- Include the difficulty field for every question
- Use "question_type": "mcq" or "question_type": "descriptive"
- The "questions" array must contain EXACTLY {total} items — no more, no fewer"""

    raw = _call_ai(prompt)
    data = _extract_json(raw)
    questions = data.get("questions", [])
    # Enforce exact count — truncate silently if AI returns too many
    return [_normalize_question_quotations(q) for q in questions[:total]]


def _evaluate_attempt(questions: List[TrainingQuestion], answers: List[dict]) -> dict:
    qa_pairs = []
    answer_map = {a["question_id"]: a["answer_text"] for a in answers}
    for seq_num, q in enumerate(questions, start=1):
        qa_pairs.append({
            "question_id": q.id,
            "question_number": seq_num,  # sequential 1..N — use this in feedback text, NOT question_id
            "question_type": q.question_type,
            "question_text": q.question_text,
            "options": q.options,
            "correct_answer": q.correct_answer,
            "user_answer": answer_map.get(q.id, ""),
        })

    total_q = len(qa_pairs)
    prompt = f"""You are evaluating a training assessment submission of {total_q} questions.

Questions and answers:
{json.dumps(qa_pairs, indent=2)}

For each question, evaluate the user's answer.

Return ONLY valid JSON (no markdown, no explanation):
{{
  "evaluations": [
    {{
      "question_id": <copy the question_id field exactly>,
      "is_correct": true,
      "ai_flag": "correct",
      "ai_explanation": "Brief explanation of why this answer is correct or wrong"
    }}
  ],
  "overall_feedback": "2-3 sentences of overall feedback on the submission",
  "score": 75.0
}}

CRITICAL RULES:
- The evaluations array MUST have exactly {total_q} entries — one per question, in the same order.
- "question_id" in each evaluation MUST be copied exactly from the question_id field in the input — do NOT substitute or invent new IDs.
- In "overall_feedback" and "ai_explanation", ALWAYS refer to questions by their "question_number" (e.g. "Question 3", "Q5"), NEVER by their "question_id" (which is an internal database ID, not a sequence number).
- ai_flag must be "correct", "wrong", or "partial" — nothing else.
- MCQ: correct only if the user's letter matches the correct_answer letter (case-insensitive). Wrong otherwise — no partial credit for MCQ.
- Descriptive (question_type = "descriptive" or "written"): partial credit allowed; compare the user's answer against the model answer and use your judgment.
- score = (correct_count + 0.5 * partial_count) / {total_q} * 100, rounded to 1 decimal.
- ai_explanation is REQUIRED for EVERY question and must NEVER be empty — this applies to correct, wrong, and partial answers alike.
- In ai_explanation, ALWAYS state the correct answer using its full text, not just the option letter. For MCQ, write out the actual option wording (e.g. 'The correct answer is "They" — used when the subject's gender is unspecified.'), NEVER just 'The correct answer is A'.
- For a wrong answer, briefly explain why the chosen answer is incorrect AND state the full correct answer text. For a correct answer, briefly confirm why it is right.
- Be constructive and specific. Reference the actual content from the question."""

    raw = _call_ai(prompt)
    return _extract_json(raw)


# ── Serialisers ──────────────────────────────────────────────────────────────

def _correct_answer_display(q) -> str:
    """Human-readable correct answer: the full option text for MCQ (not just the
    letter), or the model answer text for descriptive questions."""
    raw = (q.correct_answer or "").strip()
    if q.question_type == "mcq" and q.options:
        letter = raw[:1].upper()
        for opt in q.options:
            o = str(opt).strip()
            # options look like "A. text" or "A) text"
            if o[:1].upper() == letter and (len(o) < 2 or o[1] in ".)- "):
                return o
        return raw
    # Descriptive: strip the "Model answer:" prefix for a cleaner display
    if raw.lower().startswith("model answer:"):
        return raw[len("model answer:"):].strip()
    return raw


def _assessment_out(a: TrainingAssessment, include_questions=False, for_joiner=False, generation=None) -> dict:
    d = {
        "id": a.id,
        "title": a.title,
        "new_joiner_id": a.new_joiner_id,
        "created_by": a.created_by,
        "sme_kit_id": a.sme_kit_id,
        "source_file_ids": a.source_file_ids or [],
        "total_questions": a.total_questions,
        "mcq_count": a.mcq_count,
        "written_count": a.written_count,
        "easy_count": a.easy_count or 0,
        "medium_count": a.medium_count or 0,
        "hard_count": a.hard_count or 0,
        "pass_threshold": a.pass_threshold,
        "status": a.status,
        "attempt_request_status": a.attempt_request_status,
        "created_at": a.created_at,
        "new_joiner_name": a.new_joiner.name if a.new_joiner else None,
        "creator_name": a.creator.name if a.creator else None,
        "kit_name": a.kit.name if a.kit else None,
    }
    if include_questions:
        # Only expose the latest generation of questions (or generation passed as hint)
        all_qs = sorted(a.questions, key=lambda x: x.order_index)
        max_gen = max((q.generation or 1 for q in all_qs), default=1) if all_qs else 1
        target_gen = generation if generation is not None else max_gen
        qs = []
        for q in all_qs:
            if (q.generation or 1) != target_gen:
                continue
            qd = {
                "id": q.id,
                "assessment_id": q.assessment_id,
                "order_index": q.order_index,
                "question_type": q.question_type,
                "difficulty": q.difficulty,
                "question_text": _normalize_us_quotation_marks(q.question_text),
                "options": [_normalize_us_quotation_marks(option) for option in q.options] if isinstance(q.options, list) else q.options,
            }
            if not for_joiner:
                qd["correct_answer"] = _normalize_us_quotation_marks(q.correct_answer)
            qs.append(qd)
        d["questions"] = qs
    return d


def _attempt_out(attempt: TrainingAttempt) -> dict:
    answers_out = []
    for ans in attempt.answers:
        a = {
            "id": ans.id,
            "question_id": ans.question_id,
            "answer_text": ans.answer_text,
            "is_correct": ans.is_correct,
            "ai_flag": ans.ai_flag,
            "ai_explanation": ans.ai_explanation,
        }
        # Embed the question data so history works regardless of which generation was used
        if ans.question:
            a["question_text"] = _normalize_us_quotation_marks(ans.question.question_text)
            a["question_type"] = ans.question.question_type
            a["difficulty"] = ans.question.difficulty
            a["options"] = [_normalize_us_quotation_marks(option) for option in ans.question.options] if isinstance(ans.question.options, list) else ans.question.options
            a["order_index"] = ans.question.order_index
            a["correct_answer"] = _normalize_us_quotation_marks(ans.question.correct_answer)
            a["correct_answer_text"] = _normalize_us_quotation_marks(_correct_answer_display(ans.question))
        answers_out.append(a)
    # Sort by order_index so they display in the right sequence
    answers_out.sort(key=lambda x: x.get("order_index", 0))
    return {
        "id": attempt.id,
        "assessment_id": attempt.assessment_id,
        "user_id": attempt.user_id,
        "attempt_number": attempt.attempt_number,
        "question_generation": attempt.question_generation or 1,
        "status": attempt.status,
        "score": attempt.score,
        "passed": attempt.passed,
        "trophy_awarded": attempt.trophy_awarded,
        "ai_feedback": attempt.ai_feedback,
        "submitted_at": attempt.submitted_at,
        "evaluated_at": attempt.evaluated_at,
        "created_at": attempt.created_at,
        "answers": answers_out,
    }


# ── Routes ───────────────────────────────────────────────────────────────────

@router.post("/assessments/generate", response_model=dict)
def generate_assessment(
    payload: GenerateAssessmentRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("manager", "admin")),
):
    kit = db.query(SmeKit).filter(SmeKit.id == payload.sme_kit_id).first()
    if not kit:
        raise HTTPException(404, "SME Kit not found")

    joiner = db.query(User).filter(User.id == payload.new_joiner_id).first()
    if not joiner:
        raise HTTPException(404, "New joiner not found")

    files = db.query(SmeKitFileV2).filter(
        SmeKitFileV2.id.in_(payload.source_file_ids),
        SmeKitFileV2.sme_kit_id == payload.sme_kit_id,
    ).all()
    if not files:
        raise HTTPException(400, "No valid files found in kit for given IDs")

    content = _build_content_context(files)

    # Reject before calling AI if no file has readable content.
    # "no text extracted" / "no transcript provided" means the AI will hallucinate
    # questions about the error string rather than the actual document.
    _NO_CONTENT_MARKERS = ("no text extracted", "no transcript provided", "No content available")
    has_file_with_content = any(
        (f.transcript or "").strip() and
        not any(marker in (f.transcript or "") for marker in _NO_CONTENT_MARKERS)
        for f in files
    )
    has_real_content = has_file_with_content or (
        content.strip() and
        not any(marker in content for marker in _NO_CONTENT_MARKERS)
    )
    if not has_real_content or content.strip() == "No content available.":
        raise HTTPException(
            400,
            "The selected file has no readable text content. "
            "Please open the SME Kit, select the file, and paste the document text into the Transcript field manually. "
            "This is needed for PDFs that are scanned images or have copy-protection — the text cannot be extracted automatically."
        )

    easy_type = getattr(payload, 'easy_type', 'mcq') or 'mcq'
    medium_type = getattr(payload, 'medium_type', 'mcq') or 'mcq'
    hard_type = getattr(payload, 'hard_type', 'descriptive') or 'descriptive'
    questions_data = _generate_questions(
        content,
        payload.easy_count, easy_type,
        payload.medium_count, medium_type,
        payload.hard_count, hard_type,
        kit_name=kit.name or "",
        additional_instructions=payload.additional_instructions or "",
    )

    total = payload.easy_count + payload.medium_count + payload.hard_count
    # Count MCQ vs descriptive based on chosen types
    mcq_count = sum([
        payload.easy_count if easy_type == 'mcq' else 0,
        payload.medium_count if medium_type == 'mcq' else 0,
        payload.hard_count if hard_type == 'mcq' else 0,
    ])
    written_count = total - mcq_count

    assessment = TrainingAssessment(
        title=payload.title,
        new_joiner_id=payload.new_joiner_id,
        created_by=current_user.id,
        sme_kit_id=payload.sme_kit_id,
        source_file_ids=payload.source_file_ids,
        total_questions=total,
        mcq_count=mcq_count,
        written_count=written_count,
        easy_count=payload.easy_count,
        medium_count=payload.medium_count,
        hard_count=payload.hard_count,
        additional_instructions=(payload.additional_instructions or "").strip()[:2000] or None,
        pass_threshold=payload.pass_threshold,
        status="active",
    )
    db.add(assessment)
    db.flush()

    for qd in questions_data:
        raw_type = qd.get("question_type", "descriptive")
        # Normalise: old "written" values → "descriptive"
        if raw_type == "written":
            raw_type = "descriptive"
        q = TrainingQuestion(
            assessment_id=assessment.id,
            order_index=qd.get("order_index", 0),
            question_type=raw_type,
            difficulty=qd.get("difficulty"),
            question_text=qd.get("question_text", ""),
            options=qd.get("options"),
            correct_answer=qd.get("correct_answer"),
        )
        db.add(q)

    # Notify the new joiner that a quiz has been assigned
    db.add(Notification(
        user_id=payload.new_joiner_id,
        title=f"New Quiz Assigned: {payload.title}",
        message=f"{current_user.name} assigned you the quiz \"{payload.title}\" ({total} questions). Open AI Quizzes to take it.",
        type="quiz_assigned",
    ))

    db.commit()
    db.refresh(assessment)
    return _assessment_out(assessment, include_questions=True, for_joiner=False)


@router.post("/assessments/from-excel", response_model=dict)
async def create_assessment_from_excel(
    file: UploadFile = File(...),
    new_joiner_id: int = Form(...),
    title: str = Form(...),
    pass_threshold: int = Form(70),
    sheet: str = Form("all"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("manager", "admin")),
):
    """Import questions from an Excel file (3-sheet template) instead of AI generation."""
    import openpyxl

    joiner = db.query(User).filter(User.id == new_joiner_id).first()
    if not joiner:
        raise HTTPException(404, "New joiner not found")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(400, f"Could not read Excel file: {str(e)[:120]}")

    DESCRIPTIVE_SHEET = "Descriptive questions"
    all_sheets = wb.sheetnames
    sheets_to_read = all_sheets if sheet == "all" else [sheet]

    questions_data = []
    order = 1

    for sheet_name in sheets_to_read:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        is_descriptive = sheet_name == DESCRIPTIVE_SHEET

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            difficulty = str(row[0] or "").strip().lower()
            if difficulty not in ("easy", "medium", "hard"):
                continue

            question_text = str(row[2] or "").strip()
            if not question_text:
                continue
            question_text = _normalize_us_quotation_marks(question_text)

            if is_descriptive:
                answer_explanation = _normalize_us_quotation_marks(str(row[3] or "").strip())
                questions_data.append({
                    "order_index": order,
                    "question_type": "descriptive",
                    "difficulty": difficulty,
                    "question_text": question_text,
                    "options": None,
                    "correct_answer": f"Model answer: {answer_explanation}",
                })
            else:
                opt_a = _normalize_us_quotation_marks(str(row[3] or "").strip())
                opt_b = _normalize_us_quotation_marks(str(row[4] or "").strip())
                opt_c = _normalize_us_quotation_marks(str(row[5] or "").strip())
                opt_d = _normalize_us_quotation_marks(str(row[6] or "").strip())
                correct = str(row[7] or "").strip()
                questions_data.append({
                    "order_index": order,
                    "question_type": "mcq",
                    "difficulty": difficulty,
                    "question_text": question_text,
                    "options": [f"A. {opt_a}", f"B. {opt_b}", f"C. {opt_c}", f"D. {opt_d}"],
                    "correct_answer": correct,
                })
            order += 1

    if not questions_data:
        raise HTTPException(400, "No valid questions found in the uploaded file. Check the file format.")

    total = len(questions_data)
    mcq_count = sum(1 for q in questions_data if q["question_type"] == "mcq")
    written_count = total - mcq_count
    easy_count = sum(1 for q in questions_data if q["difficulty"] == "easy")
    medium_count = sum(1 for q in questions_data if q["difficulty"] == "medium")
    hard_count = sum(1 for q in questions_data if q["difficulty"] == "hard")

    assessment = TrainingAssessment(
        title=title.strip(),
        new_joiner_id=new_joiner_id,
        created_by=current_user.id,
        sme_kit_id=None,
        source_file_ids=[],
        total_questions=total,
        mcq_count=mcq_count,
        written_count=written_count,
        easy_count=easy_count,
        medium_count=medium_count,
        hard_count=hard_count,
        pass_threshold=pass_threshold,
        status="active",
    )
    db.add(assessment)
    db.flush()

    for qd in questions_data:
        db.add(TrainingQuestion(
            assessment_id=assessment.id,
            order_index=qd["order_index"],
            question_type=qd["question_type"],
            difficulty=qd["difficulty"],
            question_text=qd["question_text"],
            options=qd["options"],
            correct_answer=qd["correct_answer"],
        ))

    db.commit()
    db.refresh(assessment)
    d = _assessment_out(assessment, include_questions=False)
    d["attempt_count"] = 0
    d["best_score"] = None
    d["passed"] = False
    d["attempts"] = []
    return d


@router.get("/assessments", response_model=List[dict])
def list_assessments_manager(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("manager", "admin")),
):
    q = db.query(TrainingAssessment)
    if current_user.role == "manager":
        q = q.filter(TrainingAssessment.created_by == current_user.id)
    assessments = q.order_by(TrainingAssessment.created_at.desc()).all()
    result = []
    for a in assessments:
        d = _assessment_out(a, include_questions=True, for_joiner=False)
        attempts = db.query(TrainingAttempt).filter(
            TrainingAttempt.assessment_id == a.id,
        ).order_by(TrainingAttempt.attempt_number.desc()).all()
        d["attempt_count"] = len(attempts)
        d["best_score"] = max((at.score for at in attempts if at.score is not None), default=None)
        d["passed"] = any(at.passed for at in attempts)
        d["attempts"] = [_attempt_out(at) for at in attempts]
        result.append(d)
    return result


@router.get("/assessments/mine", response_model=List[dict])
def list_assessments_joiner(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("new_joiner")),
):
    assessments = (
        db.query(TrainingAssessment)
        .filter(
            TrainingAssessment.new_joiner_id == current_user.id,
            TrainingAssessment.status == "active",
        )
        .order_by(TrainingAssessment.created_at.desc())
        .all()
    )
    result = []
    for a in assessments:
        d = _assessment_out(a)
        attempts = db.query(TrainingAttempt).filter(
            TrainingAttempt.assessment_id == a.id,
            TrainingAttempt.user_id == current_user.id,
        ).all()
        evaluated = [at for at in attempts if at.status == "evaluated"]
        d["attempt_count"] = len(evaluated)
        d["best_score"] = max((at.score for at in evaluated if at.score is not None), default=None)
        d["passed"] = any(at.passed for at in evaluated)
        result.append(d)
    return result


@router.get("/assessments/{assessment_id}", response_model=dict)
def get_assessment(
    assessment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("manager", "admin", "new_joiner")),
):
    a = db.query(TrainingAssessment).filter(TrainingAssessment.id == assessment_id).first()
    if not a:
        raise HTTPException(404, "Assessment not found")
    is_joiner = current_user.role == "new_joiner"
    if is_joiner and a.new_joiner_id != current_user.id:
        raise HTTPException(403, "Not your assessment")

    # For joiners: return the question generation matching their in-progress attempt
    gen = None
    if is_joiner:
        in_progress = db.query(TrainingAttempt).filter(
            TrainingAttempt.assessment_id == assessment_id,
            TrainingAttempt.user_id == current_user.id,
            TrainingAttempt.status == "in_progress",
        ).first()
        if in_progress:
            gen = in_progress.question_generation or 1

    result = _assessment_out(a, include_questions=True, for_joiner=is_joiner, generation=gen)
    result["retake_wait_seconds"] = 0
    result["retake_available_at"] = None
    if is_joiner:
        latest_failed = db.query(TrainingAttempt).filter(
            TrainingAttempt.assessment_id == assessment_id,
            TrainingAttempt.user_id == current_user.id,
            TrainingAttempt.status == "evaluated",
            TrainingAttempt.passed.is_(False),
            TrainingAttempt.submitted_at.isnot(None),
        ).order_by(TrainingAttempt.submitted_at.desc()).first()
        if latest_failed:
            available_at = latest_failed.submitted_at + timedelta(minutes=15)
            wait_seconds = max(0, int((available_at - datetime.utcnow()).total_seconds()))
            result["retake_wait_seconds"] = wait_seconds
            result["retake_available_at"] = available_at.isoformat()
    return result


@router.post("/assessments/{assessment_id}/start", response_model=dict)
def start_attempt(
    assessment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("new_joiner")),
):
    from models import TrainingQuestion as TQ
    a = db.query(TrainingAssessment).filter(TrainingAssessment.id == assessment_id).first()
    if not a or a.new_joiner_id != current_user.id:
        raise HTTPException(404, "Assessment not found")
    if a.status != "active":
        raise HTTPException(400, "Assessment is not active")

    in_progress = db.query(TrainingAttempt).filter(
        TrainingAttempt.assessment_id == assessment_id,
        TrainingAttempt.user_id == current_user.id,
        TrainingAttempt.status == "in_progress",
    ).first()
    if in_progress:
        return _attempt_out(in_progress)

    # Enforce 3-attempt limit
    evaluated = db.query(TrainingAttempt).filter(
        TrainingAttempt.assessment_id == assessment_id,
        TrainingAttempt.user_id == current_user.id,
        TrainingAttempt.status == "evaluated",
    ).all()
    has_passed = any(at.passed for at in evaluated)
    failed_count = sum(1 for at in evaluated if not at.passed)
    if not has_passed and failed_count >= 3:
        if a.attempt_request_status != "approved":
            raise HTTPException(403, "Attempt limit reached. Request a new attempt from your manager.")
        # Consume the approval — reset so the joiner must request again if they fail
        a.attempt_request_status = None
        db.flush()

    latest_failed = max(
        (attempt for attempt in evaluated if not attempt.passed and attempt.submitted_at),
        key=lambda attempt: attempt.submitted_at,
        default=None,
    )
    if latest_failed:
        available_at = latest_failed.submitted_at + timedelta(minutes=15)
        now = datetime.utcnow()
        if now < available_at:
            wait_seconds = max(1, int((available_at - now).total_seconds()))
            wait_minutes = (wait_seconds + 59) // 60
            raise HTTPException(
                status_code=429,
                detail={
                    "code": "retake_cooldown",
                    "message": f"Retake available in {wait_minutes} minute(s).",
                    "wait_seconds": wait_seconds,
                    "available_at": available_at.isoformat(),
                },
            )

    attempt_number = db.query(TrainingAttempt).filter(
        TrainingAttempt.assessment_id == assessment_id,
        TrainingAttempt.user_id == current_user.id,
    ).count() + 1

    # Determine which question generation to use for this attempt
    existing_gens = db.query(TQ.generation).filter(TQ.assessment_id == assessment_id).distinct().all()
    max_existing_gen = max((g[0] or 1 for g in existing_gens), default=1)

    if attempt_number == 1:
        # First attempt — use existing question set (generation 1)
        question_generation = 1
    elif a.sme_kit_id is None:
        # Excel-imported assessment — no SME Kit to regenerate from.
        # Shuffle the latest generation's questions into a new generation so the
        # order is different each attempt (prevents memorisation of position).
        import random
        new_generation = max_existing_gen + 1
        source_qs = [q for q in a.questions if (q.generation or 1) == max_existing_gen]
        shuffled = list(source_qs)
        random.shuffle(shuffled)
        for new_order, orig_q in enumerate(shuffled, start=1):
            db.add(TrainingQuestion(
                assessment_id=a.id,
                order_index=new_order,
                question_type=orig_q.question_type,
                difficulty=orig_q.difficulty,
                question_text=orig_q.question_text,
                options=orig_q.options,
                correct_answer=orig_q.correct_answer,
                generation=new_generation,
            ))
        db.commit()
        question_generation = new_generation
    else:
        # AI-generated assessment — regenerate a fresh question set from SME Kit content
        new_generation = max_existing_gen + 1
        try:
            files = db.query(SmeKitFileV2).filter(
                SmeKitFileV2.id.in_(a.source_file_ids or []),
                SmeKitFileV2.sme_kit_id == a.sme_kit_id,
            ).all()
            content = _build_content_context(files)
            easy_type = "mcq"
            medium_type = "mcq"
            hard_type = "descriptive"
            gen1_qs = [q for q in a.questions if (q.generation or 1) == 1]
            if gen1_qs:
                easy_qs = [q for q in gen1_qs if (q.difficulty or "").lower() == "easy"]
                medium_qs = [q for q in gen1_qs if (q.difficulty or "").lower() == "medium"]
                hard_qs = [q for q in gen1_qs if (q.difficulty or "").lower() == "hard"]
                if easy_qs:
                    easy_type = easy_qs[0].question_type
                if medium_qs:
                    medium_type = medium_qs[0].question_type
                if hard_qs:
                    hard_type = hard_qs[0].question_type
            questions_data = _generate_questions(
                content,
                a.easy_count or 0, easy_type,
                a.medium_count or 0, medium_type,
                a.hard_count or 0, hard_type,
                kit_name=a.kit.name if a.kit else "",
                additional_instructions=a.additional_instructions or "",
            )
            for qd in questions_data:
                raw_type = qd.get("question_type", "descriptive")
                if raw_type == "written":
                    raw_type = "descriptive"
                db.add(TrainingQuestion(
                    assessment_id=a.id,
                    order_index=qd.get("order_index", 0),
                    question_type=raw_type,
                    difficulty=qd.get("difficulty"),
                    question_text=qd.get("question_text", ""),
                    options=qd.get("options"),
                    correct_answer=qd.get("correct_answer"),
                    generation=new_generation,
                ))
            db.commit()
            question_generation = new_generation
        except Exception as regen_err:
            # AI regeneration failed — raise so the learner retries rather than
            # silently serving the same questions in a shuffled order
            raise HTTPException(
                503,
                "Could not generate new questions from the SME Kit. Please try again in a moment.",
            ) from regen_err

    attempt = TrainingAttempt(
        assessment_id=assessment_id,
        user_id=current_user.id,
        attempt_number=attempt_number,
        question_generation=question_generation,
        status="in_progress",
    )
    db.add(attempt)
    db.commit()
    db.refresh(attempt)
    return _attempt_out(attempt)


@router.post("/assessments/{assessment_id}/submit", response_model=dict)
def submit_attempt(
    assessment_id: int,
    payload: SubmitAttemptRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("new_joiner")),
):
    a = db.query(TrainingAssessment).filter(TrainingAssessment.id == assessment_id).first()
    if not a or a.new_joiner_id != current_user.id:
        raise HTTPException(404, "Assessment not found")

    attempt = db.query(TrainingAttempt).filter(
        TrainingAttempt.assessment_id == assessment_id,
        TrainingAttempt.user_id == current_user.id,
        TrainingAttempt.status == "in_progress",
    ).first()
    if not attempt:
        # Check if already submitted/evaluated (e.g. client timeout but server completed)
        recent = db.query(TrainingAttempt).filter(
            TrainingAttempt.assessment_id == assessment_id,
            TrainingAttempt.user_id == current_user.id,
            TrainingAttempt.status.in_(["submitted", "evaluated"]),
        ).order_by(TrainingAttempt.id.desc()).first()
        if recent:
            return _attempt_out(recent)
        raise HTTPException(400, "No in-progress attempt found. Start the assessment first.")

    attempt.status = "submitted"
    attempt.submitted_at = datetime.utcnow()
    db.commit()

    answers_input = payload.answers

    # Look up questions directly by the IDs submitted by the frontend.
    # This is generation-agnostic and avoids mismatches if question_generation
    # doesn't perfectly align with what was displayed to the user.
    submitted_ids = [ans.get("question_id") for ans in answers_input if ans.get("question_id")]
    questions_by_id = {q.id: q for q in a.questions}
    questions = sorted(
        [questions_by_id[qid] for qid in submitted_ids if qid in questions_by_id],
        key=lambda x: x.order_index,
    )
    # Fallback: if no submitted IDs match (shouldn't happen), use generation filter
    if not questions:
        attempt_gen = attempt.question_generation or 1
        questions = sorted(
            [q for q in a.questions if (q.generation or 1) == attempt_gen],
            key=lambda x: x.order_index,
        )

    try:
        evaluation = _evaluate_attempt(questions, answers_input)
        evals = {e["question_id"]: e for e in evaluation.get("evaluations", [])}
        score = float(evaluation.get("score", 0.0))
        overall_feedback = evaluation.get("overall_feedback", "")
    except Exception as exc:
        # AI evaluation failed — auto-grade MCQ questions at least so the
        # learner gets a real score rather than 0
        answer_map = {ans.get("question_id"): ans.get("answer_text", "") for ans in answers_input}
        correct_count = 0
        total_q = len(questions)
        evals = {}
        for q in questions:
            user_ans = (answer_map.get(q.id) or "").strip().upper()
            correct = (q.correct_answer or "").strip().upper()
            correct_text = _correct_answer_display(q)
            if q.question_type == "mcq" and user_ans and correct and user_ans == correct:
                correct_count += 1
                evals[q.id] = {
                    "question_id": q.id, "is_correct": True, "ai_flag": "correct",
                    "ai_explanation": f"Correct. The right answer is: {correct_text}",
                }
            elif q.question_type == "mcq":
                evals[q.id] = {
                    "question_id": q.id, "is_correct": False, "ai_flag": "wrong",
                    "ai_explanation": f"The correct answer is: {correct_text}",
                }
            else:
                # Descriptive — cannot auto-grade; show the model answer for reference
                evals[q.id] = {
                    "question_id": q.id, "is_correct": None, "ai_flag": "partial",
                    "ai_explanation": f"This answer could not be auto-evaluated. Model answer for reference: {correct_text}",
                }
        score = round((correct_count / total_q) * 100, 1) if total_q > 0 else 0.0
        overall_feedback = "Descriptive answers could not be AI-evaluated at this time. MCQ answers have been auto-graded."

    for ans_data in answers_input:
        qid = ans_data.get("question_id")
        ev = evals.get(qid, {})
        answer = TrainingAnswer(
            attempt_id=attempt.id,
            question_id=qid,
            answer_text=ans_data.get("answer_text", ""),
            is_correct=ev.get("is_correct"),
            ai_flag=ev.get("ai_flag"),
            ai_explanation=ev.get("ai_explanation"),
        )
        db.add(answer)

    passed = score >= a.pass_threshold
    attempt.score = score
    attempt.passed = passed
    attempt.trophy_awarded = score >= 90
    attempt.ai_feedback = {"overall": overall_feedback}
    attempt.status = "evaluated"
    attempt.evaluated_at = datetime.utcnow()

    # Notify the manager who created the quiz that the new joiner completed it
    if a.created_by:
        result_label = "passed" if passed else "did not pass"
        db.add(Notification(
            user_id=a.created_by,
            title=f"Quiz Completed: {a.title}",
            message=f"{current_user.name} completed \"{a.title}\" and {result_label} with a score of {round(score)}/100.",
            type="quiz_completed",
        ))

    db.commit()
    db.refresh(attempt)
    return _attempt_out(attempt)


@router.get("/assessments/{assessment_id}/attempts", response_model=List[dict])
def list_attempts(
    assessment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("manager", "admin", "new_joiner")),
):
    a = db.query(TrainingAssessment).filter(TrainingAssessment.id == assessment_id).first()
    if not a:
        raise HTTPException(404, "Assessment not found")
    if current_user.role == "new_joiner" and a.new_joiner_id != current_user.id:
        raise HTTPException(403, "Not your assessment")

    attempts = db.query(TrainingAttempt).filter(
        TrainingAttempt.assessment_id == assessment_id,
    ).order_by(TrainingAttempt.attempt_number.desc()).all()
    return [_attempt_out(at) for at in attempts]


@router.get("/attempts/{attempt_id}", response_model=dict)
def get_attempt(
    attempt_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("manager", "admin", "new_joiner")),
):
    attempt = db.query(TrainingAttempt).filter(TrainingAttempt.id == attempt_id).first()
    if not attempt:
        raise HTTPException(404, "Attempt not found")
    if current_user.role == "new_joiner" and attempt.user_id != current_user.id:
        raise HTTPException(403, "Not your attempt")
    return _attempt_out(attempt)


@router.post("/assessments/{assessment_id}/request-attempt")
def request_new_attempt(
    assessment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("new_joiner")),
):
    """New joiner requests a new attempt after exhausting the 3-attempt limit."""
    a = db.query(TrainingAssessment).filter(
        TrainingAssessment.id == assessment_id,
        TrainingAssessment.new_joiner_id == current_user.id,
    ).first()
    if not a:
        raise HTTPException(404, "Assessment not found")

    evaluated = db.query(TrainingAttempt).filter(
        TrainingAttempt.assessment_id == assessment_id,
        TrainingAttempt.user_id == current_user.id,
        TrainingAttempt.status == "evaluated",
    ).all()
    has_passed = any(at.passed for at in evaluated)
    if has_passed:
        raise HTTPException(400, "You have already passed this assessment")
    if len(evaluated) < 3:
        raise HTTPException(400, "You have not yet used all 3 attempts")

    # Mark request as pending on the assessment
    a.attempt_request_status = "pending"

    # Notify the manager
    if current_user.manager_id:
        notif = Notification(
            user_id=current_user.manager_id,
            title="New Attempt Request",
            message=f"{current_user.name} has used all 3 attempts on '{a.title}' without passing and is requesting a new attempt.",
            type="warning",
        )
        db.add(notif)
    db.commit()
    return {"ok": True, "message": "Request sent to your manager."}


@router.post("/assessments/{assessment_id}/approve-attempt")
def approve_attempt_request(
    assessment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("manager", "admin")),
):
    """Manager approves a new joiner's request for an extra attempt."""
    a = db.query(TrainingAssessment).filter(TrainingAssessment.id == assessment_id).first()
    if not a:
        raise HTTPException(404, "Assessment not found")
    if current_user.role == "manager" and a.created_by != current_user.id:
        raise HTTPException(403, "Not your assessment")

    a.attempt_request_status = "approved"

    if a.new_joiner_id:
        db.add(Notification(
            user_id=a.new_joiner_id,
            title="Attempt Request Approved",
            message=f"Your manager approved a new attempt for '{a.title}'. You can now start a new attempt.",
            type="info",
        ))
    db.commit()
    return {"ok": True}


@router.post("/assessments/{assessment_id}/reject-attempt")
def reject_attempt_request(
    assessment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("manager", "admin")),
):
    """Manager rejects a new joiner's request for an extra attempt."""
    a = db.query(TrainingAssessment).filter(TrainingAssessment.id == assessment_id).first()
    if not a:
        raise HTTPException(404, "Assessment not found")
    if current_user.role == "manager" and a.created_by != current_user.id:
        raise HTTPException(403, "Not your assessment")

    a.attempt_request_status = "rejected"

    if a.new_joiner_id:
        db.add(Notification(
            user_id=a.new_joiner_id,
            title="Attempt Request Rejected",
            message=f"Your manager has rejected your request for a new attempt on '{a.title}'. Please speak to your manager for more information.",
            type="warning",
        ))
    db.commit()
    return {"ok": True}
